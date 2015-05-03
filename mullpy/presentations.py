# This Python file uses the following encoding: utf-8
# !/usr/local/bin/python3.4
####################################################
# <Copyright (C) 2012, 2013, 2014, 2015 Yeray Alvarez Romero>
# This file is part of MULLPY.
####################################################
import re
import itertools

import numpy as np

from mullpy.auxiliar import AutoVivification

####################################################


class Presentation:
    """
    Class where are defined all graphics presentation mode of the classifiers information.
    """

    def general_paint_method(self, parameters, context, classifier_name=""):
        """
        General mode to paint graphic with parameters as args. Parameters are constructed in their respective function.
        """
        import matplotlib.pyplot as plt
        import matplotlib

        plt.rc('text', usetex=True)
        plt.rc('font', family='serif')
        # from matplotlib.ticker import MultipleLocator, FormatStrFormatter

        markers = ["o", "v", ">", "<", 'H', 'p', 'D', 'd', '^']
        limits = {"roc": {"x": [-0.1, 1.1], "y": [-0.1, 1.1]},
                  "classes_error": {"x": [1.97, parameters["num_classes"] + 0.03], "y": [-0.001, .85]},
                  "instances_error": {"x": [-0.01, 6.03], "y": [-0.001, .5]}}
        #"learning":{"x":[max(map(max,parameters["axis_x"]))/2,max(map(max,parameters["axis_x"]))],
        # "y":[-0.0001,max(map(max,parameters["axis_y"]))/10]}}
        ####
        if context["interactive"]:
            plt.ion()
        fig = plt.figure()
        ax = fig.add_subplot(111)
        plt.xlabel(parameters["xlabel"], fontsize=30)
        plt.ylabel(parameters["ylabel"], fontsize=30)
        plt.title(parameters["title"], fontsize=40)
        ###
        values = range(len(parameters["label"]))
        jet = cm = plt.get_cmap('jet')
        cNorm = matplotlib.colors.Normalize(vmin=0, vmax=values[-1])
        scalarMap = matplotlib.cm.ScalarMappable(norm=cNorm, cmap=jet)

        for i in range(len(parameters["label"])):
            if parameters["paint_kind"] == "roc":
                ax.plot(parameters["axis_x"][i], parameters["axis_y"][i], label=parameters["label"][i],
                        color=scalarMap.to_rgba(values[i]))
                ax.plot([0.0, 1.0], [0.0, 1.0], linestyle='--', color='k', lw=0.05)
            elif parameters["paint_kind"] == "classes_error" or parameters["paint_kind"] == "instances_error":
                ax.plot(parameters["axis_x"][i], parameters["axis_y"][i], label=parameters["label"][i],
                        color=scalarMap.to_rgba(values[i]), marker=markers[i % len(markers)], markersize=10)
        legend = ax.legend(loc=parameters["loc"], prop={'size': 12})
        for t in legend.get_texts():
            t.set_fontsize('large')
        for l in legend.get_lines():
            l.set_linewidth(0.7)
            #add limits to the x and y axis
        if parameters["paint_kind"] == "roc":
            ax.set_xlim(limits[parameters["paint_kind"]]["x"][0], limits[parameters["paint_kind"]]["x"][1])
            ax.set_ylim(limits[parameters["paint_kind"]]["y"][0], limits[parameters["paint_kind"]]["y"][1])
        elif parameters["paint_kind"] == "classes_error" or parameters["paint_kind"] == "instances_error":
            xmin, xmax = ax.get_xlim()
            ymin, ymax = ax.get_ylim()
            ax.set_xlim(xmin=limits[parameters["paint_kind"]]["x"][0], xmax=limits[parameters["paint_kind"]]["x"][1])
            #ax.set_ylim(ymin=0.0, ymax=np.max(parameters["axis_y"]) + 0.05 * np.max(parameters["axis_y"]))
            ax.set_ylim(ymin=0.0, ymax=parameters["ymax"])
            if parameters["paint_kind"] == "classes_error":
                plt.xticks(range(1, parameters["num_classes"] + 1), fontsize=30)
            elif parameters["paint_kind"] == "instances_error":
                plt.xticks(range(7), fontsize=30)
            elif parameters["paint_kind"] == "roc":
                plt.xticks(range(0.0, 1.0, 0.1), fontsize=30)
                ###
                #ax.yaxis.grid(True,'major',linewidth=0.1)
        plt.tight_layout()  # Adapt the elements to the limits of the figure
        if context["interactive"]:
            plt.draw()

        if parameters["paint_kind"] == "roc":
            parameters["file_name"] = context["pattern_kind"] + "_" + parameters["paint_kind"] + "_" + \
                                      classifier_name + context["result_name"] + "." + context["graphics"]["extension"]
        else:
            parameters["file_name"] = context["pattern_kind"] + "_" + parameters["paint_kind"] + "_" + \
                                      classifier_name + "_" + context["result_name"] + "." + context["graphics"][
                                          "extension"]
        print(parameters["path"] + parameters["file_name"])
        plt.savefig(parameters["path"] + parameters["file_name"])

    #####################################################
    @staticmethod
    def identify_cross_validation_number(x):
        """
        Return the k_fold number of a cross_validation member
        :param x: The classifier_name
        :return:
        """
        #First number is the only search method to identify the kfold number
        if x.count("_") > 1:
            return int(x[x.find("_")+1:x.find("_", x.find("_")+1)])
        elif x.count("_") == 1:
            return int(x[x.find("_")+1:])
        else:
            raise Exception("No pattern for identifying cross validation members known")

    #####################################################
    @staticmethod
    def identify_cross_validation_members(classifier_name):
        #First number is the only search method to identify the kfold number
        first_number = 1

        classifier_name = classifier_name.split("+")
        if first_number:
            final_name = ""
            for name in classifier_name:
                res = re.search(r'_[0-9]+', name)
                final_name += name[:res.start()] + name[res.end():]
                if name is not classifier_name[-1]:
                    final_name += "+"
            return final_name

    #####################################################

    def cross_validation(self, context, stats, order_criteria):
        summary_stats = AutoVivification()

        if context["results"]["to_file"]["show_only_ensembles"]:
            classifier_list = context["ensemble_list"]
        else:
            classifier_list = context["classifiers"].keys()

        for classifier_name in classifier_list:
            new_name = self.identify_cross_validation_members(classifier_name)
            for measure in order_criteria:
                value = np.mean([stats[classifier_name][measure]])

                if new_name in summary_stats and measure in summary_stats[new_name]:
                    #TODO: review the comparison method. Compare every classifier inside an ensemble
                    summary_stats[new_name][measure].append(value)
                else:
                    summary_stats[new_name][measure] = [value]
        return summary_stats

    #####################################################
    @staticmethod
    def write_meta_learning_patterns(context, classifier_list, info, pattern_kind, file):
        classifier_name_example = context["classifier_list"][0]
        len_classes = len(context["classifiers"][classifier_name_example]["classes_names"])
        len_inputs = len(context["patterns"].patterns[classifier_name_example][pattern_kind][0]) - len_classes

        for classifier_name in classifier_list:
            # for class_text in context["classifiers"][classifier_name]["classes_names"][1]:
            class_text = context["classifiers"][classifier_name]["classes_names"][1]
            file.write("@FEATURE {0}\n".format(classifier_name + "_" + class_text))

        for i in range(len(info[classifier_name_example]["continuous_outputs"][pattern_kind])):
            # file.write("\t".join([str(y) for x in classifier_list for y in info[x]["continuous_outputs"][
            # pattern_kind][i]]))
            file.write("\t".join([str(info[x]["continuous_outputs"][pattern_kind][i][1]) for x in classifier_list]))
            file.write("\t")
            file.write("\t".join([str(x) for x in
                                  context["patterns"].patterns[classifier_name_example][pattern_kind][i][len_inputs:]]))
            file.write("\n")
        file.close()

    #####################################################
    @staticmethod
    def create_meta_learning_patterns(folder, classifier_list, info, context):
        """
        Create a meta learning pattern from the outputs of the classifiers.
        Each feature name correspond to a classifier name, so the order is not important anymore.
        Each instance is a feature fusion of many classifiers
        :param file_descriptor: The file_descriptor
        :param classifier_list: List of classifiers names to work with
        :param info: The info structure
        :param pattern_kind: The pattern kind to work on
        """
        from mullpy.auxiliar import create_output_file

        pattern_kind = context["pattern_kind"]
        cross_validation = context["results"]["to_file"]["create_meta_learning_patterns"]["cross_validation"]
        k_folds = list(set([Presentation.identify_cross_validation_number(x) for x in classifier_list]))

        if cross_validation:
            for fold in k_folds:
                file_name = "{0}_{1}.pat".format(pattern_kind, str(fold))
                file = create_output_file(folder, file_name, "overwrite", "y")
                k_fold_list = [x for x in classifier_list if Presentation.identify_cross_validation_number(x) == fold]
                Presentation.write_meta_learning_patterns(context, k_fold_list, info, pattern_kind, file)
        else:
            file_name = "{0}_{1}.pat".format(pattern_kind, context["result_name"])
            file = create_output_file(folder, file_name, "overwrite", "y")
            Presentation.write_meta_learning_patterns(context, classifier_list, info, pattern_kind, file)

    #####################################################
    @staticmethod
    def _get_short_name(classifier_name):
        if classifier_name.find("_") != -1:
            return classifier_name[:classifier_name.find("_")]
        else:
            return classifier_name

    #####################################################

    @staticmethod
    def sort_classifier_list_by_order_criteria(classifier_list, stats, order_criteria):
        return sorted(classifier_list,
                      key=lambda x: np.sum([sum(stats[x][y]) if type(stats[x][y]) == list else stats[x][y]
                                            for y in order_criteria]), reverse=True)

    #####################################################
    @staticmethod
    def get_ordered_list_to_write_file(classifier_list, stats, order_criteria, context):
        cl_amount = context["results"]["to_file"]["retain_best_classifiers_in_file"]["amount"]
        if context["results"]["to_file"]["retain_best_classifiers_in_file"]["activate"]:
            if context["results"]["to_file"]["retain_best_classifiers_in_file"]["per_type"]:
                temp_ordered_list = Presentation.sort_classifier_list_by_order_criteria(classifier_list,
                                                                                        stats,
                                                                                        order_criteria)

                list_classifiers_kind = AutoVivification()
                for classifier_name in temp_ordered_list:
                    short_name = Presentation._get_short_name(classifier_name)
                    if short_name not in list_classifiers_kind:
                        list_classifiers_kind[short_name] = 0.

                ordered_list = []
                for classifier_name in temp_ordered_list:
                    short_name = Presentation._get_short_name(classifier_name)
                    if list_classifiers_kind[short_name] < (cl_amount // len(list_classifiers_kind.keys())):
                        ordered_list.append(classifier_name)
                        list_classifiers_kind[short_name] += 1

            else:
                ordered_list = Presentation.sort_classifier_list_by_order_criteria(classifier_list,
                                                                                   stats,
                                                                                   order_criteria)
        else:
            ordered_list = Presentation.sort_classifier_list_by_order_criteria(classifier_list,
                                                                               stats,
                                                                               order_criteria)
        return ordered_list

    #####################################################
    @staticmethod
    def write_stats_in_file(file, classifier_list, stats, order_criteria, context):
        """
        :param file: The file_descriptor
        :param classifier_list: List of classifiers names to be printed
        :param stats: Structure from statistics or a transformation from statistics which contain stats of classifiers
        :param order_criteria: A list which contains the names of the measure that will be used as a criteria to order
        :return: Write to the file descriptor

        -The order of appearance of the classifiers depends on the sum of the measures which comes in order_criteria,
        but there is no resume of the total value. It is preferable to see all the information
        -The first measure to write into the file has to be in the same order as comes in order_criteria
        """
        measures_names = order_criteria
        file.write("%s" % "".join(["\t"] * 10))

        for measure in order_criteria + [x for x in measures_names if x not in order_criteria]:
            file.write("\t{0}".format(measure))
        file.write("\n")

        ordered_list = Presentation().get_ordered_list_to_write_file(classifier_list,
                                                                     stats,
                                                                     order_criteria,
                                                                     context)
        for i, classifier_name in enumerate(ordered_list):
            file.write("{0}:\t".format(classifier_name))
            for measure in measures_names:
                if type(stats[classifier_name][measure]) == list and len(stats[classifier_name][measure]) > 0:
                    for value in stats[classifier_name][measure]:
                        if value is stats[classifier_name][measure][-1]:
                            file.write("%.5f" % value)
                        else:
                            file.write("%.5f," % value)
                else:
                    file.write("\t%.5f" % stats[classifier_name][measure])

                if measure is not measures_names[-1]:
                    file.write("\t")
            file.write("\n")

    #####################################################
    @staticmethod
    def merge_summaries(summary_recovered, summary):
        """
        Modify summary to add the new measures from summary_recovered
        :param summary_recovered:
        :param summary:
        :return:
        """
        for classifier_name in summary_recovered:
            for measure in summary_recovered[classifier_name]:
                summary[classifier_name][measure] = summary_recovered[classifier_name][measure]

    #####################################################

    def check_create_det_recover(self, context, stats, folder, output_file_name, order_criteria):
        """
        Check and create cross_validation values and write summary structure.
        Create the output_file.
        Determines the list to iterate over.
        Recover the info from a file if sequentially generation is on execution
        :param context:
        :param stats:
        :param folder:
        :param output_file_name:
        :param order_criteria:
        :return:
        """
        from mullpy.auxiliar import create_output_file
        from mullpy.auxiliar import summary_stats_from_file

        summary_recovered = AutoVivification()
        if context["results"]["to_file"]["retain_best_classifiers_in_file"]["activate"]:
            summary_recovered = summary_stats_from_file(folder, output_file_name)
            file_descriptor = create_output_file(folder, output_file_name, "overwrite", "y")
        else:
            file_descriptor = create_output_file(folder, output_file_name, "append", "y")

        if context["results"]["to_file"]["cross_validation"]:
            summary = self.cross_validation(context, stats, order_criteria)
            classifier_list = list(summary.keys())
        elif context["results"]["to_file"]["show_only_ensembles"]:
            classifier_list = context["ensemble_list"]
            summary = stats
        else:
            classifier_list = context["classifier_list"] + context["ensemble_list"]
            summary = stats

        if context["results"]["to_file"]["retain_best_classifiers_in_file"]["activate"]:
            self.merge_summaries(summary_recovered, summary)
            classifier_list = list(summary.keys())

        return file_descriptor, summary, classifier_list

    #####################################################

    def plain_text(self, context, stats, info):
        #Due to the reconstruction of the context from each classifier file, we have to reconstruct the path here.
        set_name = context["classifiers"][context["classifier_list"][0]]["set"]
        folder = "{0}{1}/{2}/".format(context["general_path"], "results", set_name)
        output_file_name = context["pattern_kind"] + "_" + context["result_name"] + ".txt"

        order_criteria = context["results"]["to_file"]["order_criteria"]

        if "create_meta_learning_patterns" in context["results"]["to_file"] and context["results"]["to_file"][
                "create_meta_learning_patterns"]["activate"]:
            folder = "{0}{1}/{2}/".format(context["general_path"], "patterns", "meta_" + set_name)
            self.create_meta_learning_patterns(folder, context["classifier_list"], info, context)
        else:
            file_descriptor, summary, classifier_list = self.check_create_det_recover(context,
                                                                                      stats,
                                                                                      folder,
                                                                                      output_file_name,
                                                                                      order_criteria)

            self.write_stats_in_file(file_descriptor, classifier_list, summary, order_criteria, context)
            file_descriptor.close()

    #####################################################

    def excel(self, context, stats, info):
        """
        Construct an excel table with the information contained in the stats parameter as Statistic Class.
        """
        from openpyxl import Workbook
        from openpyxl.cell import get_column_letter
        #Style for numbers printed in table
        #style = Style()
        #style.num_format_str='0.000'
        book = Workbook()
        book.remove_sheet(book.get_active_sheet())

        sheet_general = book.create_sheet()
        sheet_general.title = "General"

        #Information relative to the Process in the first page
        #[TO-DO] Incorporate all the context without limited depth
        column = 1
        for row, characteristic in enumerate(context):
            sheet_general.cell('%s%s' % (get_column_letter(column), row + 1)).value = characteristic
            if context[characteristic].__class__.__name__ == "str" or context[
                characteristic].__class__.__name__ == "int" or context[characteristic].__class__.__name__ == "float":
                sheet_general.cell('%s%s' % (get_column_letter(column + 1), row + 1)).value = context[characteristic]
            elif context[characteristic].__class__.__name__ == "list":
                if len(context[characteristic]):
                    sheet_general.cell('%s%s' % (get_column_letter(column + 1), row + 1)).value = \
                        [str(x) for x in context[characteristic]]
                    #####################################################
        sheet_stats = book.create_sheet()
        sheet_stats.title = "Stats"
        #####################################################
        #Write the headers of the information type
        #####################################################
        classifier_base = context["classifier_list"][0]
        metric_list = sorted(stats[classifier_base].keys())
        row = 1
        sheet_stats.cell('%s%s' % (get_column_letter(1), 1)).value = "All classes Error"
        for column in range(2, len(metric_list) + 2):
            sheet_stats.cell('%s%s' % (get_column_letter(column), row)).value = metric_list[column]
            # for column in range(2, len(["error", "error_fn", "error_fp"]) * 2 + 2, 2):
            #     sheet_stats.cell('%s%s' % (get_column_letter(column), row)).value = \
            #         str(["error", "error_fn", "error_fp"][int((column - 1) / 2)])
            # if context["results"]["to_file"]["measures"]["std"]:
            #     for column in range(3, len(context["deviations_texts"]) * 2 + 3, 2):
            #         sheet_stats.cell('%s%s' % (get_column_letter(column), row)).value = \
            #             str(context["deviations_texts"][int((column - 2) / 2)])
            # old_column = column + 1
            # for column in range(old_column, old_column + len(context["goodness_texts"])):
            #     sheet_stats.cell('%s%s' % (get_column_letter(column), row)).value = \
            #         str(context["goodness_texts"][column - old_column])
            # old_column = column + 1
            # if context["results"]["to_file"]["measures"]["rms"]:
            #     temp_list = ["rms_" + x for x in stats[classifier_base]["rms"].keys()]
            #     for column, value in zip(range(old_column, old_column + len(temp_list)), temp_list):
            #         sheet_stats.cell('%s%s' % (get_column_letter(column), row)).value = str(value)
            #####################################################
            #Classifiers error values and goodness
            #####################################################
        row = 2
        column = 1
        for row, classifier_name in zip(range(row, len(stats) + 2),
                                        [x for x in sorted(context["classifier_list"],
                                                           key=lambda y: stats[y]['E'])]):
            sheet_stats.cell('%s%s' % (get_column_letter(column), row)).value = classifier_name
            for column_iteration in range(column + 1, len(["error", "error_fn", "error_fp"]) * 2 + 2, 2):
                sheet_stats.cell('%s%s' % (get_column_letter(column_iteration), row)).value = \
                    str(stats[classifier_name][["error", "error_fn", "error_fp"][int((column_iteration - 1) / 2)]])
            if context["results"]["to_file"]["measures"]["std"]:
                for column_iteration in range(column + 2, len(context["deviations_texts"]) * 2 + 3, 2):
                    sheet_stats.cell('%s%s' % (get_column_letter(column_iteration), row)).value = \
                        str(stats[classifier_name][context["deviations_texts"][int((column_iteration - 2) / 2)]])
            old_column_iteration = column_iteration + 1
            for column_iteration in range(old_column_iteration, old_column_iteration + len(context["goodness_texts"])):
                sheet_stats.cell('%s%s' % (get_column_letter(column_iteration), row)).value = \
                    str(stats[classifier_name][context["goodness_texts"][column_iteration - old_column_iteration]])
            old_column_iteration = column_iteration + 1
            if context["results"]["to_file"]["measures"]["rms"]:
                for column_iteration, pattern in zip(
                        range(old_column_iteration, old_column_iteration + len(stats[classifier_base]["rms"].keys())),
                        stats[classifier_base]["rms"].keys()):
                    sheet_stats.cell('%s%s' % (get_column_letter(column_iteration), row)).value = \
                        str(stats[classifier_name]["rms"][pattern])
                    ####################################################
                    #SPECIFIC INFORMATION BY CLASS
                    #####################################################
        classes_text = list(set([context["classifiers"][x]["classes_names"][y] for x in stats.keys()
                                 for y in range(len(context["classifiers"][x]["classes_names"]))]))

        row = (len(stats) + 8)
        for file_numbers, class_text in enumerate(classes_text):
            for column in range(2, len(["error", "error_fn", "error_fp"]) * 2 + 2, 2):
                sheet_stats.cell('%s%s' % (get_column_letter(column), row)).value = \
                    str(["error", "error_fn", "error_fp"][int((column - 1) / 2)])
            if context["results"]["to_file"]["measures"]["std"]:
                for column in range(3, len(context["deviations_texts"]) * 2 + 3, 2):
                    sheet_stats.cell('%s%s' % (get_column_letter(column), row)).value = \
                        str(context["deviations_texts"][int((column - 2) / 2)])
            old_column = column + 1
            for column in range(old_column, old_column + len(context["goodness_texts"])):
                sheet_stats.cell('%s%s' % (get_column_letter(column), row)).value = \
                    str(context["goodness_texts"][column - old_column])
            column = 2
            sheet_stats.cell('%s%s' % (get_column_letter(column - 1), row - 1)).value = class_text
            for row, classifier_name in zip(range(row, row + len(stats) + 2),
                                            [x for x in sorted(list(context["classifier_list"]),
                                                               key=lambda y: stats[y]['E'])
                                             if
                                             class_text in context["classifiers"][x]["classes_names"]]):
                sheet_stats.cell('%s%s' % (get_column_letter(column - 1), row)).value = str(classifier_name)
                for column_iteration in range(2, len(["error", "error_fn", "error_fp"]) * 2 + 1, 2):
                    sheet_stats.cell('%s%s' % (get_column_letter(column_iteration), row)).value = \
                        str(stats[classifier_name][class_text][
                            ["error", "error_fn", "error_fp"][int((column_iteration - 1) / 2)]])
                if context["results"]["to_file"]["measures"]["std"]:
                    for column_iteration in range(3, len(context["deviations_texts"]) * 2 + 3, 2):
                        sheet_stats.cell('%s%s' % (get_column_letter(column_iteration), row)).value = \
                            str(stats[classifier_name][class_text][
                                context["deviations_texts"][int((column_iteration - 2) / 2)]])
                old_column_iteration = column_iteration + 1
                for column_iteration in range(old_column_iteration,
                                              old_column_iteration + len(context["goodness_texts"])):
                    sheet_stats.cell('%s%s' % (get_column_letter(column_iteration), row)).value = \
                        str(stats[classifier_name][class_text][
                            context["goodness_texts"][column_iteration - old_column_iteration]])

            row += (np.sum([1 for x in context["classifiers"].keys() if
                            class_text in context["classifiers"][x]["classes_names"]]) + 5)
            #####################################################
            #DIVERSIY METRICS IN A NEW SHEET
            #####################################################
        if context["results"]["to_file"]["diversity_study"]["activate"]:
            sheet_diversity = book.create_sheet()
            sheet_diversity.title = "Diversity"

            old_column = 1
            row = 1

            if sum(context["results"]["to_file"]["diversity_study"]["pairwise_diversity"].values()):
                for correlation_kind in [x for x in
                                         context["results"]["to_file"]["diversity_study"]["pairwise_diversity"].keys()
                                         if
                                         context["results"]["to_file"]["diversity_study"]["pairwise_diversity"][
                                             x] == 1]:

                    ordered_list = [x for x in sorted(context["classifier_list"],
                                                      key=lambda y: stats[y]["pairwise_diversity"][correlation_kind][
                                                          "mean"])]

                    sheet_diversity.cell('%s%s' % (get_column_letter(old_column), row)).value = correlation_kind
                    for column, classifier_name in zip(range(old_column + 1, old_column + 1 + len(ordered_list)),
                                                       ordered_list):
                        sheet_diversity.cell('%s%s' % (get_column_letter(column), row)).value = classifier_name
                    for row, classifier_name in zip(range(row + 1, row + 1 + len(ordered_list)), ordered_list):
                        sheet_diversity.cell('%s%s' % (get_column_letter(old_column), row)).value = classifier_name

                    old_column = 2
                    row -= len(context["classifier_list"]) - 1
                    for row, classifier_1 in zip(range(row, row + len(ordered_list) + 2), ordered_list):
                        for column, classifier_2 in zip(range(old_column, old_column + len(ordered_list)),
                                                        ordered_list):
                            try:
                                sheet_diversity.cell('%s%s' % (get_column_letter(column), row)).value = \
                                    str(stats[classifier_1]["pairwise_diversity"][correlation_kind][classifier_2])
                            except:
                                print(len(stats[classifier_1]["pairwise_diversity"][correlation_kind].keys()))
                        old_column = 2

                    row += 1  # To see clearly when the classifiers diversity ends
                    for statistic in context["results"]["to_file"]["diversity_study"]["statistics"]:
                        row += 1
                        sheet_diversity.cell('%s%s' % (get_column_letter(1), row)).value = statistic
                        for column, classifier_1 in zip(range(2, len(ordered_list) + 2), ordered_list):
                            sheet_diversity.cell('%s%s' % (get_column_letter(column), row)).value = \
                                str(stats[classifier_1]["pairwise_diversity"][correlation_kind][statistic])

                    row += 3
                    old_column = 1
            if sum(context["results"]["to_file"]["diversity_study"]["non_pairwise_diversity"].values()):
                for diversity_kind in [k for k, v in
                                       context["results"]["to_file"]["diversity_study"][
                                           "non_pairwise_diversity"].items()
                                       if v == 1]:
                    sheet_diversity.cell('%s%s' % (get_column_letter(1), row)).value = diversity_kind
                    for classifiers_combination in sorted(
                            [x for x in stats.keys() if stats[x] != {}],
                            key=lambda y: stats[y][diversity_kind], reverse=True):
                        row += 1
                        sheet_diversity.cell('%s%s' % (get_column_letter(1), row)).value = classifiers_combination
                        sheet_diversity.cell('%s%s' % (get_column_letter(2), row)).value = \
                            str(stats[classifiers_combination][diversity_kind])

                    row += 2

                    #####################################################
                    #OUTPUTS OF THE CLASSIFIERS IN A NEW SHEET
                    #####################################################
        if context["results"]["to_file"]["outputs"]:
            sheet_outputs = book.create_sheet()
            sheet_outputs.title = "Outputs"
            outputs_kind = context["outputs_kind"]
            pattern_kind = context["pattern_kind"]

            column = 2
            for counter, classifier_name in enumerate(info.keys()):
                classes = context["classifiers"][classifier_name]["classes_names"]
                sheet_outputs.cell('%s%s' % (get_column_letter(column), 1)).value = classifier_name

                for row in range(2, len(context["patterns"].patterns[classifier_name][pattern_kind]) + 3):
                    if row > 2:
                        sheet_outputs.cell('%s%s' % (get_column_letter(1), row)).value = row - 2
                    for col_iteration in range(len(classes)):
                        if row == 2:
                            sheet_outputs.cell('%s%s' % (get_column_letter(col_iteration + 2), row)).value = classes[
                                col_iteration]
                        else:
                            sheet_outputs.cell('%s%s' % (get_column_letter(column + col_iteration), row)).value = \
                                str(info[classifier_name][outputs_kind][pattern_kind][row - 3][col_iteration])

                column += len(classes) + 1
                ##SAVE
        book.save(
            context["classifiers"][context["classifier_list"][0]]["paths"]["results"] +
            context["pattern_kind"] + "_" + context["result_name"] + ".xlsx")

    #######################################################################
    def weights_file(self, context, classifier_name):
        """
        Create a file with the information of the classifier, like weigths or thresholds.
        [TO-DO] Generalize the method to print all classifiers info structurally
        """
        f = open(
            context["classifiers"][classifier_name]["paths"]["results"] + classifier_name + "_pesos.txt",
            "w")
        for i in range(len(context["classifiers_instances"][classifier_name].wi)):
            for j in range(len(context["classifiers_instances"][classifier_name].wi[i])):
                f.write(repr(context["classifiers_instances"][classifier_name].wi[i][j]) + " ")
            f.write("\n")
        f.write("\n")
        f.write("\n")
        for i in range(len(context["classifiers_instances"][classifier_name].wo)):
            for j in range(len(context["classifiers_instances"][classifier_name].wo[i])):
                f.write(repr(context["classifiers_instances"][classifier_name].wo[i][j]) + " ")
            f.write("\n")
        f.write("UMBRALES: \n")
        for threshold in context["classifiers"][classifier_name]["thresholds"]["value"]:
            f.write(str(threshold) + "\n")
        f.close()

    #######################################################################
    def roc_curve(self, context, stats):
        """
        Specific function where are defined the parameters to be passed to the  general paint function method to paint the ROCs.
        """
        for classifier_name in context["ensemble_list"]:
            graphic_parameters = AutoVivification()
            self._latex_string_convert(context, classifier_name)
            graphic_parameters["xlabel"] = context["graphics"]["roc"]["xlabel"]
            graphic_parameters["ylabel"] = context["graphics"]["roc"]["ylabel"]
            graphic_parameters["title"] = context["classifiers"][classifier_name]["name_to_show"]

            graphic_parameters["axis_x"] = []
            classes_texts = context["classifiers"][classifier_name]["classes_names"]

            for class_text in classes_texts:
                graphic_parameters["axis_x"].append(stats[classifier_name][class_text]['FPR'])
                #graphic_parameters["axis_x"].append(classifiers_information.info[classifier_name]['FPR'])

            graphic_parameters["axis_y"] = []
            for class_text in classes_texts:
                graphic_parameters["axis_y"].append(stats[classifier_name][class_text]['TPR'])
                #graphic_parameters["axis_y"].append(classifiers_information.info[classifier_name]['TPR'])

            graphic_parameters["num_classes"] = len(classes_texts)
            graphic_parameters["paint_kind"] = "roc"
            graphic_parameters["loc"] = "lower right"
            graphic_parameters["label"] = [class_text for class_text in classes_texts]
            #graphic_parameters["label"].append(name_to_show)

            graphic_parameters["path"] = context["classifiers"][classifier_name]["paths"]["results"]

            self.general_paint_method(graphic_parameters, context, classifier_name)

    #######################################################################
    def paint_classes_error(self, context, Information):
        """
        Define the parameters to be passed to the function of general paint to paint the classes_error Information.info.
        """
        graphic_parameters = AutoVivification()
        graphic_parameters["xlabel"] = context["graphics"]["classes_error"]["xlabel"]
        graphic_parameters["ylabel"] = context["graphics"]["classes_error"]["ylabel"]
        graphic_parameters["title"] = context["graphics"]["classes_error"]["title"]

        classifier_list = sorted(context["ensemble_list"])

        graphic_parameters["axis_x"] = [range(1, len(
            Information.info[context["classifier_list"][0]]["selection_errors"]) + 1)] * len(classifier_list)

        graphic_parameters["axis_y"] = []
        for classifier_name in classifier_list:
            #Ensembles and classifiers
            graphic_parameters["axis_y"].append(Information.info[classifier_name]["selection_errors"][:])

        #print(print(classifier_name, Information.info[classifier_name]["selection_errors"]))
        graphic_parameters["loc"] = "best"

        #for classifier_name in classifier_list:
        #    self._latex_string_convert(context, classifier_name)
        graphic_parameters["label"] = [context["classifiers"][x]["name_to_show"] for x in
                                       classifier_list]

        graphic_parameters["path"] = context["classifiers"][context["classifier_list"][0]]["paths"]["results"]

        graphic_parameters["paint_kind"] = "classes_error"
        graphic_parameters["num_classes"] = len(Information.info[context["classifier_list"][0]]["selection_errors"])

        self.general_paint_method(graphic_parameters, context)

    #######################################################################

    def paint_instances_error(self, context, Information):
        """
        Define the parameters to be passed to the function of general paint to paint the instances error model.
        """
        graphic_parameters = AutoVivification()
        graphic_parameters["xlabel"] = context["graphics"]["instances_error"]["xlabel"]
        graphic_parameters["ylabel"] = context["graphics"]["instances_error"]["ylabel"]
        graphic_parameters["title"] = context["graphics"]["instances_error"]["title"]

        #classifier_list = context["classifier_list"] + context["ensemble_list"]
        classifier_list = sorted(context["ensemble_list"])

        graphic_parameters["axis_y"] = []
        for classifier_name in classifier_list:
            graphic_parameters["axis_y"].append(Information.info[classifier_name]["selection_errors"])

        graphic_parameters["axis_x"] = [range(len(graphic_parameters["axis_y"][0]))] * len(graphic_parameters["axis_y"])
        graphic_parameters["loc"] = "best"

        #for classifier_name in classifier_list:
        #self._latex_string_convert(context, classifier_name)

        graphic_parameters["label"] = [context["classifiers"][x]["name_to_show"] for x in classifier_list]
        graphic_parameters["xmax"] = context["graphics"]["instances_error"]["xmax"]
        graphic_parameters["ymax"] = context["graphics"]["instances_error"]["ymax"]

        graphic_parameters["num_classes"] = len(context["classifiers"][context["classifier_list"][0]]["classes_names"])
        graphic_parameters["path"] = context["classifiers"][context["classifier_list"][0]]["paths"]["results"]

        graphic_parameters["paint_kind"] = "instances_error"
        self.general_paint_method(graphic_parameters, context)

    #######################################################################

    def validation_bars(self, context, info):
        """
        Method to construct all the information about validation bars like:
            -Plain bars containing:
                -Errors by kind of error
                -Errors by classes
                -Comparison between the best modules to each ensemble
            -3D Bars containing:
                -Errors by kind of error
                -Errors by classes
        """
        import matplotlib.pyplot as plt
        import matplotlib

        parameters = AutoVivification()
        #Order the list by error, taking as a second criterion the number of classifiers presents in the Ensemble
        names_list = [x[0] for x in itertools.chain(sorted(info.items(), key=lambda y: (y[1]['E'], y[0].count("+"))))]
        #Modules only or modules and ensembles in the same graph
        if context["results"]["validation_bars"]["no_modules"]:
            names_list = [x for x in names_list if x not in context["classifier_list"]]
            #Select the bests or the worsts
        if context["results"]["validation_bars"]["worst_bars"]:
            names_list = names_list[len(names_list) - context["results"]["validation_bars"]["number_or_bars"]:]
        else:
            names_list = names_list[:context["results"]["validation_bars"]["number_or_bars"]]
            #COLORS Scalarmap
        values = range(len(names_list))
        jet = cm = plt.get_cmap('prism')
        cNorm = matplotlib.colors.Normalize(vmin=0, vmax=values[-1])
        scalarMap = matplotlib.cm.ScalarMappable(norm=cNorm, cmap=jet)
        ###GRAPH CONFIGURATION
        fig = plt.figure()
        plt.xlabel(context["graphics"]["validation_bars"]["xlabel"])
        plt.ylabel(context["graphics"]["validation_bars"]["ylabel"])
        plt.title(context["graphics"]["validation_bars"]["title"])
        #Get the 'name to show' labels gived by user
        labels = [context["classifiers"][x]["name_to_show"] if x in context["classifiers"] else
                  context["classifiers"][x]["name_to_show"] for x in names_list]
        #################
        if sum(context["results"]["validation_bars"]["error_type"].values()):
            legend = 0
            colors = ["r", "b", "g", "y", "m", '#07dcca', '#9cfd4e', 'k']
            if context["results"]["validation_bars"]["error_type"]["class_text"]:
                width = 0.8 / len(context["classifiers"][context["classifier_list"][0]][
                    "classes_names"])
                for class_text, posi in zip(
                        context["classifiers"][context["classifier_list"][0]][
                            "classes_names"], range(len(
                                context["classifiers"][context["classifier_list"][0]][
                                    "classes_names"]))):
                    parameters["axis_x"] = range(len(names_list))
                    parameters["axis_x"] = [x + posi * width for x in parameters["axis_x"]]
                    parameters["axis_y"] = [info[classifier_name][class_text]["e"] for classifier_name in names_list]
                    plt.bar(parameters["axis_x"], parameters["axis_y"], color=colors[posi], width=width,
                            label=class_text)
                    #Write the numeric value over the bar
                    #for i in range(len(parameters["axis_y"])):
                    #plt.text(parameters["axis_x"][i],parameters["axis_y"][i],s="%.3f"%(parameters["axis_y"][i]),
                    # fontsize=6,ha='left',rotation=80,va='bottom')
            elif context["results"]["validation_bars"]["error_type"]["error"]:
                width = 0.8 / len(["error", "error_fn", "error_fp"])
                for error, posi in zip(["error", "error_fn", "error_fp"],
                                       range(len(["error", "error_fn", "error_fp"]))):
                    parameters["axis_x"] = range(len(names_list))
                    parameters["axis_x"] = [x + posi * width for x in parameters["axis_x"]]
                    parameters["axis_y"] = [info[classifier_name][error] for classifier_name in names_list]
                    bar = plt.bar(parameters["axis_x"], parameters["axis_y"], color=colors[posi], width=width,
                                  label=error)
                    for i in range(len(parameters["axis_y"])): plt.text(parameters["axis_x"][i],
                                                                        parameters["axis_y"][i],
                                                                        s="%.3f" % (parameters["axis_y"][i]),
                                                                        fontsize=6, ha='left', rotation=80, va='bottom')
                    #############################
                    #Best classifier vs Ensemble#
                    #############################
            elif context["results"]["validation_bars"]["error_type"]["best_module"]:
                width = 0.3
                ##########
                #Print the Ensemble error bar
                ##########
                parameters["axis_x"] = range(len(names_list))
                parameters["axis_y"] = [info[classifier_name]["e"] for classifier_name in names_list]
                plt.bar(parameters["axis_x"], parameters["axis_y"], color='k', width=width)
                #Write the numeric value over the bar
                #for i in range(len(parameters["axis_y"])):
                #plt.text(parameters["axis_x"][i],parameters["axis_y"][i],s="%.3f"%(parameters["axis_y"][i]),
                # fontsize=6,ha='left',rotation=80,va='bottom')
                ########################################
                #Paint the best module vs Ensemble
                ########################################
                parameters["axis_x"] = [x - width for x in range(len(names_list))]
                parameters["axis_y"] = np.zeros(len(names_list))

                best_modules_list = []
                colors_list = []
                colors_module = {}
                for ensemble_name in [x for x in names_list if x in context["ensemble_list"]]:
                    parameters["axis_y"][names_list.index(ensemble_name)] = sorted(
                        [info[x]["e"] for x in context["classifiers"][ensemble_name]["classifiers"]])[0]
                    temp = sorted([x for x in context["classifiers"][ensemble_name]["classifiers"]],
                                  key=lambda y: info[y]['E'])[0]
                    best_modules_list.append(context["classifiers"][temp]["name_to_show"])
                    if context["classifiers"][temp]["name_to_show"] not in colors_module.keys():
                        colors_module[context["classifiers"][temp]["name_to_show"]] = len(best_modules_list)
                        colors_list.append(colors_module[context["classifiers"][temp]["name_to_show"]])
                        #Color scale
                values = range(len(colors_list))
                jet = cm = plt.get_cmap('prism')
                cNorm = matplotlib.colors.Normalize(vmin=0, vmax=values[-1])
                scalarMap = matplotlib.cm.ScalarMappable(norm=cNorm, cmap=jet)

                plt.bar(parameters["axis_x"], parameters["axis_y"], color=scalarMap.to_rgba(range(len(values))),
                        width=width)
                #Write the numeric value over the bar
                #for i in range(len(parameters["axis_y"])):
                #plt.text(parameters["axis_x"][i],parameters["axis_y"][i],s="%.3f"%(parameters["axis_y"][i]),fontsize=6,ha='left',rotation=80,va='bottom')
            #########
            #Legends#
            #########
            legend_texts = []
            if context["results"]["validation_bars"]["error_type"]["class_text"]:
                legend_texts = context["classifiers"][context["classifier_list"][0]][
                    "classes_names"]
            if context["results"]["validation_bars"]["error_type"]["error"]:
                legend_texts.extend(["error", "error_fn", "error_fp"])
            if context["results"]["validation_bars"]["error_type"]["best_module"]:
                legend_texts.append("Error")
                legend_texts.extend(list(set(best_modules_list)))

            legend = plt.legend(legend_texts, loc="upper left")
            ########################
            #Common characteristics#
            ########################
            xmin, xmax = plt.xlim()
            ymin, ymax = plt.ylim()
            plt.xlim(xmin=-.5, xmax=len(names_list))
            if ymax > 0.7:
                plt.ylim(ymin=-0.001, ymax=ymax)
            else:
                plt.ylim(ymin=-0.001, ymax=0.7)
            if legend:
                for t in legend.get_texts():
                    t.set_fontsize('small')
            plt.xticks(range(len(labels)), labels, fontsize=6, rotation=20, ha='right')
        else:
            ###########
            #3D FIGURE
            ###########

            ax = fig.add_subplot(111, projection='3d')
            ax.view_init(elev=45, azim=-50) #perspective and angle of the figure
            ####
            z = np.zeros(len(labels))
            if context["results"]["validation_bars_3D"]["error_type"]["class_text"]:
            #########################################################
            #class_text error, first bars in the graph
            #########################################################
                for class_text, posi in zip(
                        context["classifiers"][context["classifier_list"][0]][
                            "classes_names"], range(len(
                                context["classifiers"][context["classifier_list"][0]][
                                    "classes_names"]))):
                    parameters["axis_x"] = range(len(names_list))
                    parameters["axis_y"] = [info[classifier_name][class_text]["e"] for classifier_name in names_list]
                    z = [posi] * len(parameters["axis_y"])
                    ax.bar(parameters["axis_x"], parameters["axis_y"], zs=z, zdir='y', color=scalarMap.to_rgba(values),
                           width=0.8, alpha=0.4, label=class_text)
                z = [len(context["classifiers"][context["classifier_list"][0]][
                    "classes_names"])] * len(parameters["axis_y"])
                #########################################################
            #Bars by kind of error
            #########################################################
            if context["results"]["validation_bars_3D"]["error_type"]["error"]:
                for error, posi in zip(reversed(["error", "error_fn", "error_fp"]),
                                       range(len(["error", "error_fn", "error_fp"]))):
                    parameters["axis_x"] = range(len(names_list))
                    parameters["axis_y"] = [info[classifier_name][error] for classifier_name in names_list]
                    if context["results"]["validation_bars_3D"]["error_type"]["class_text"]:
                        z = [z[0] + posi] * len(parameters["axis_y"])
                    else:
                        z = [z[0] + posi] * len(parameters["axis_y"])
                    if error != 'E':
                        ax.bar(parameters["axis_x"], parameters["axis_y"], zs=z, zdir='y',
                               color=scalarMap.to_rgba(values), width=0.8, alpha=0.4, label=error)
                z = [z[0] - 1] * len(parameters["axis_y"])
                #########################################################
            #Global error of the classifiers
            #########################################################
            parameters["axis_x"] = range(len(names_list))
            parameters["axis_y"] = [info[classifier_name]["e"] for classifier_name in names_list]
            ax.bar(parameters["axis_x"], parameters["axis_y"], zs=z, zdir='y', color=scalarMap.to_rgba(values),
                   width=0.8, alpha=0.6)
            z = [z[0] + 1.] * len(parameters["axis_y"])
            #########################################################
            #Best module vs Ensemble
            #########################################################
            if context["results"]["validation_bars_3D"]["error_type"]["best_module"]:
                parameters["axis_x"] = range(len(names_list))
                parameters["axis_y"] = np.zeros(len(names_list))
                if context["ensembling"]:
                    for ensemble_name in [x for x in names_list if x in context["ensemble_list"]]:
                        for classifier_name in context["classifiers"][ensemble_name]["classifiers"]:
                            parameters["axis_y"][names_list.index(ensemble_name)] += info[classifier_name]["e"]
                        parameters["axis_y"][names_list.index(ensemble_name)] /= len(
                            context["classifiers"][ensemble_name]["classifiers"])
                    rects = ax.bar(parameters["axis_x"], parameters["axis_y"], zs=z, zdir='y',
                                   color=scalarMap.to_rgba(values), width=0.8, alpha=1.0)
                    #############################################
                    #Axis labels
                    ##########
                    #ZTICKS
            zmin, zmax = ax.get_zlim()
            if zmax > 0.7:
                ax.set_zlim(0.0, zmax)
            else:
                ax.set_zlim(0.0, 0.7)
                #list_temp=[float("%.2f"%(x)) for x in (list(set(sorted(np.append(ax.get_zticks(),(parameters[
                # "axis_y"]))))))]
                #ax.set_zticks([list_temp[i] for i in range(len(list_temp)) if (i>1 and ((list_temp[i]-
                # list_temp[i-1])>0.01))])
            #for x in ax.zaxis.get_major_ticks():x.label.set_fontsize(4)
            #Yticks
            y_texts = []
            if context["results"]["validation_bars_3D"]["error_type"]["error"] and \
                    context["results"]["validation_bars_3D"]["error_type"]["class_text"]:
                y_texts = ["Error de " + x for x in
                           context["classifiers"][context["classifier_list"][0]][
                               "classes_names"]]
                y_texts += [x.replace("e", "Error ") for x in list(reversed(["error", "error_fn", "error_fp"]))]
            elif context["results"]["validation_bars_3D"]["error_type"]["error"]:
                y_texts = [x.replace("e", "Error ") for x in list(reversed(["error", "error_fn", "error_fp"]))]
            elif context["results"]["validation_bars_3D"]["error_type"]["class_text"]:
                y_texts = ["Error de " + x for x in
                           context["classifiers"][context["classifier_list"][0]][
                               "classes_names"]] + ["Error"]
            if context["ensembling"] and context["results"]["validation_bars_3D"]["error_type"]["best_module"]:
                if not context["results"]["validation_bars_3D"]["error"] and not \
                    context["results"]["validation_bars_3D"]["class_text"]:
                    y_texts += ['Error']
                y_texts += ['Media Error Modulos']
            plt.yticks(range(len(y_texts)), y_texts, fontsize=6, ha='left')
            #
            ax.zaxis.grid(True)
            ##########
            plt.xticks(range(len(labels)), labels, fontsize=4, rotation=0, ha='right')
            ##########
        plt.tight_layout()
        if context["interactive"]:
            plt.show()
        else:
            plt.savefig(context["classifiers"][context["classifier_list"][0]]["paths"]["results"] + context[
                "pattern_kind"] + "_bars_" + context["result_name"] + context["graphics"]["extension"])

    #######################################################################

    def error_vs_rms(self, context, info):
        """
        Paint a comparison bar graphic wich contain disggregated information in rms vs E.
        Used as stability comparison between rms classifiers function based as MLP.
        Print in console the mean of each kind of error and their standard deviation.
        """
        import matplotlib.pyplot as plt

        parameters = AutoVivification()
        #Graphic configuration
        fig = plt.figure()
        colors = ["m", 'g', "r", "b", "y", "#9cfd4e", '#07dcca', 'k']
        plt.xlabel(context["graphics"]["error_vs_rms"]["xlabel"])
        plt.ylabel(context["graphics"]["error_vs_rms"]["ylabel"])
        plt.title(context["graphics"]["error_vs_rms"]["title"])

        names_list = [x[0] for x in itertools.chain(
            sorted(info.items(), key=lambda y: y[1]['E']))] #Sort classifiers passed by Error magnitude

        labels = [context["classifiers"][x]["name_to_show"] for x in
                  context["classifier_list"]]  #Get the 'name to show' labels gived by user
        #################
        if context["results"]["rms_vs_E"]["class_text"]:
            width = .8 / (len(context["classifiers"][context["classifier_list"][0]][
                "classes_names"]) + 2)
        else:
            width = 0.4
        parameters["axis_x"]["e"] = [x for x in range(len(labels))]
        parameters["axis_x"]["rms"] = [x + width for x in range(len(labels))]
        parameters["axis_y"]["e"] = [info[classifier_name]["e"] for classifier_name in names_list]
        parameters["axis_y"]["rms"] = [info[classifier_name]["rms"] for classifier_name in names_list]
        plt.bar(parameters["axis_x"]["e"], parameters["axis_y"]["e"], color='g', width=width, label="e")
        plt.bar(parameters["axis_x"]["rms"], parameters["axis_y"]["rms"], color='m', width=width, label="rms")
        if context["results"]["rms_vs_E"]["Show_std_mean"]:
            bbox_props = dict(boxstyle="round,pad=0.3", ec="k", fc="white")
            plt.text((len(labels) / 2.) - 1., parameters["axis_y"]["e"][len(parameters["axis_y"]["e"]) / 2] + 0.1, \
                     r"$\bar{X}=$" + str(
                         np.around(np.mean(parameters["axis_y"]["e"]), decimals=5)) + "\n" + r"$\sigma=$" + \
                     str(np.around(np.std(parameters["axis_y"]["e"]), decimals=5)), style='italic', fontsize=12,
                     bbox=bbox_props, va='top')
        if context["results"]["rms_vs_E"]["class_text"]:
            for class_text, posi in zip(
                    context["classifiers"][context["classifier_list"][0]][
                        "classes_names"], range(2, len(context["classifiers"][context["classifier_list"][0]][
                        "classes_names"]) + 2)):
                parameters["axis_x"] = range(len(info))
                parameters["axis_x"] = [x + posi * width for x in parameters["axis_x"]]
                parameters["axis_y"] = [info[classifier_name][class_text]["e"] for classifier_name in names_list]
                plt.bar(parameters["axis_x"], parameters["axis_y"], color=colors[posi], width=width, label=class_text)
        legend = plt.legend(loc="upper left")
        for t in legend.get_texts():
            t.set_fontsize('small')
            #Common characteristics
        xmin, xmax = plt.xlim()
        ymin, ymax = plt.ylim()
        plt.xlim(xmin=-.5, xmax=len(labels))
        if ymax > 0.4:
            plt.ylim(ymin=-0.001, ymax=ymax)
        else:
            plt.ylim(ymin=-0.001, ymax=0.4)
        if legend:
            for t in legend.get_texts():
                t.set_fontsize('small')
        plt.xticks(range(len(labels)), labels, fontsize=6, rotation=20, ha='right')
        #####
        plt.tight_layout()
        if context["interactive"]:
            plt.show()
        else:
            plt.savefig(context[context["classifier_list"][0]]["paths"]["results"] + context[
                "pattern_kind"] + "_rms_vs_E_" + labels[0] + context["graphics"]["extension"], transparent=True)

    ################################################################

    def learning_graphic(self, context, classifier_name, learning_accumulated_error, test_accumulated_error):
        import matplotlib.pyplot as plt

        plt.clf()
        ax = plt.subplot(111)
        ax.plot(range(len(learning_accumulated_error)), learning_accumulated_error, color='r')
        ax.plot(range(len(test_accumulated_error)), test_accumulated_error, color='g')
        plt.legend(["Training RMSE", "Validation RMSE"], loc="best")
        file_name = context["classifiers"][classifier_name]["paths"]["results"] + "_learning_" + \
                    classifier_name + ".pdf"
        plt.savefig(file_name)

    ################################################################

    def learning_instability(self, context, stats):
        """
        Paint a boxplot
        """
        import matplotlib.pyplot as plt
        from matplotlib import pylab
        #Graphic configuration
        plt.figure()
        #colors = ["m", 'g', "r", "b", "y", "#9cfd4e", '#07dcca', 'k']
        plt.xlabel(context["graphics"]["learning_instability"]["xlabel"])
        plt.ylabel(context["graphics"]["learning_instability"]["ylabel"])
        plt.title(context["graphics"]["learning_instability"]["title"])

        #################
        #Unique names
        unique_names = list(set([x[:re.search(r'[A-Za-z]+[0-9]', x).end()] for x in context["classifier_list"]]))
        data_list = AutoVivification()

        for classifier_name in unique_names:
            data_list[classifier_name] = []

        counter = 0
        for classifier_name in context["classifier_list"]:
            for name in unique_names:
                if name in classifier_name:
                    if stats.measures[classifier_name]["e"].__class__.__name__ == "AutoVivification":
                        print(classifier_name)
                        counter += 1
                    data_list[name].append(stats.measures[classifier_name]["e"])

        print(counter)
        data = [data_list[x] for x in data_list]

        pylab.boxplot(data)
        plt.xticks(range(1, len(unique_names) + 1), unique_names, fontsize=10, rotation=20, ha='right')
        #####
        plt.tight_layout()
        if context["interactive"]:
            plt.show()
        else:
            plt.savefig(context["classifiers"][context["classifier_list"][0]]["paths"]["results"] + context[
                "pattern_kind"] + "_" + "learning_instabilities" + "." + context["graphics"]["extension"])

    ################################################################

    def _latex_string_convert(self, context, classifier_name):
        """
        Convert the string built on the generation of the ensemble combination to the latex format. Better print quality acquired
        """
        if classifier_name in context["classifier_list"]:
            temp = context["classifiers"][classifier_name]["name_to_show"]
            pos = re.search(r'\d+', temp)

            temp = temp[:pos.start()] + "$_{%d}$" % (int(temp[pos.start():pos.end()]))
            context["classifiers"][classifier_name]["name_to_show"] = temp

        elif classifier_name in context["ensemble_list"]:
            ranges = [range(i, x + 1) for i in range(1, len(context["classifiers"][classifier_name]["classifiers"]) + 1)
                      for x in range(i + 2, len(context["classifiers"][classifier_name]["classifiers"]) + 1)]

            classifiers_string = "".join(
                [x[:x.find("-")] for x in context["classifiers"][classifier_name]["classifiers"]])
            numbers_string = ""
            numbers = [int(x) for x in re.findall(r'\d*', classifiers_string) if (len(x))]

            if numbers in ranges:  # Replace subsequent enumerations
                for range_i in ranges:
                    if range_i == numbers:
                        numbers_string += str(range_i[0]) + "-" + str(range_i[-1])
            else:
                for i in range(len(numbers)):
                    if i == len(numbers) - 1:
                        numbers_string += str(numbers[i])
                    else:
                        numbers_string += str(numbers[i]) + "+"

            if len(numbers_string) > 20:
                numbers_string = ""

            context["classifiers"][classifier_name]["name_to_show"] = r"%s$_{" % context["classifiers"][
                classifier_name]["name_to_show"] + numbers_string + "}$"

    ################################################################

    def scatter(self, context, info):
        import matplotlib.pyplot as plt
        import matplotlib

        for classifier_name in context["classifiers"].keys():
            self._latex_string_convert(context, classifier_name)
        plt.rc('text', usetex=True)
        plt.rc('font', family='serif')

        names_list = [x[0] for x in itertools.chain(sorted(info.items(), key=lambda y: (y[1]['E'], y[0].count("+")))) if
                      x[0] in context["ensemble_list"]]
        #b_list = ["BPN1+BPN2+BPN3+BPN4", "BPN1+BPN2+BPN3+BPN4+BPN5+BPN6+BPN7+BPN8", "BPN5+BPN6+BPN7+BPN8"]
        #b_list=["H1+H2+H4+H6","H1+H2+H3+H4", "H1+H2+H3+H4+H5+H6+H7+H8", "H5+H6+H7+H8"]
        #b_list = names_list
        b_list = []
        markers = ['+', 'D', 'h', 'o', '^']
        colors = ['r', 'm', 'g', 'c', 'y']
        #for i in b_list:
        #    names_list.pop(names_list.index(i))

        x = np.array([info[classifier_name]["e"] for classifier_name in names_list])
        y = np.zeros(len(x))
        if context["results"]["scatter"]["mean_module"]:
            for i in range(len(names_list)):
                y[i] = np.mean(
                    [info[z]["e"] for z in context["classifiers"][names_list[i]]["classifiers"]])
        elif context["results"]["scatter"]["best_module"]:
            for i in range(len(names_list)):
                y[i] = \
                    sorted(
                        [info[z]["e"] for z in context["classifiers"][names_list[i]]["classifiers"]])[0]
                ##
        if context["results"]["scatter"]["mean_module"] and b_list != []:
            for i, best in zip(range(len(b_list)), b_list):
                plt.scatter(info[best]["e"], np.mean(
                    [info[z]["e"] for z in context["classifiers"][best]["classifiers"]]), s=100,
                            c=colors[i], marker=markers[i], label=context["classifiers"][best]["name_to_show"])
        elif context["results"]["scatter"]["best_module"] and b_list != []:
            for i, best in zip(range(len(b_list)), b_list):
                plt.scatter(info[best]["e"],
                            sorted([info[z]["e"] for z in context["classifiers"][best]["classifiers"]])[
                                0], s=100, c=colors[i], marker=markers[i],
                            label=context["classifiers"][best]["name_to_show"])

        #Scatter best module
        best_module = sorted(context["classifier_list"], key=lambda y: (info[y]['E']))[0]
        plt.scatter(info[best_module]["e"], info[best_module]["e"], s=100, c="y", marker="^",
                    label=context["classifiers"][best_module]["name_to_show"])

        #All the combinations
        worst_x = [x[i] for i in range(len(x)) if info[best_module]["e"] <= x[i]]
        worst_y = [y[i] for i in range(len(y)) if info[best_module]["e"] <= x[i]]
        best_x = [x[i] for i in range(len(x)) if info[best_module]["e"] > x[i]]
        best_y = [y[i] for i in range(len(y)) if info[best_module]["e"] > x[i]]
        plt.scatter(worst_x, worst_y, c='k', s=5, marker=".")
        plt.scatter(best_x, best_y, c='k', s=5, marker=".") #label="Best than %s"%(best_module)

        #Plot the 5 best ensembles
        for i, best_ensembles in enumerate(sorted(context["ensemble_list"], key=lambda y: (info[y]['E']))[:5]):
            x_x = info[best_ensembles]["e"]
            y_y = np.mean([info[z]["e"] for z in context["classifiers"][best_ensembles]["classifiers"]])
            plt.scatter(x_x, y_y, s=100, c=colors[i], marker=markers[i],
                        label=context["classifiers"][best_ensembles]["name_to_show"])

        #Plot the all members ensemble
        name = sorted(context["ensemble_list"], key=lambda y: y.count("+"))[0]
        plt.scatter(info[name]["e"], np.mean([info[z]["e"] for z in context["classifiers"][name]["classifiers"]]),
                    s=100, c="r", marker="^", label="All members")

        #Plot the middle-line and the area
        plt.plot([0.0, context["graphics"]["scatter"]["xmax"]], [0.0, context["graphics"]["scatter"]["xmax"]],
                 linestyle='--', color='k', lw=0.3)
        plt.plot([info[best_module]["e"], info[best_module]["e"]], [0.0, context["graphics"]["scatter"]["ymax"]],
                 linestyle='--', color='k', lw=0.1)
        plt.gca().add_patch(
            matplotlib.patches.Rectangle((0.0, 0.0), info[best_module]["e"], 1.0, ec='k', fill=False, fc='k',
                                         ls='dotted', lw=0.3, hatch="\\"))

        plt.xlabel("Ensemble error", fontsize=25)
        if context["results"]["scatter"]["mean_module"]:
            plt.ylabel("Mean modules error", fontsize=25)
            plt.title("Ensemble error vs Mean modules error", fontsize=25)
            context["result_name"] += "mean_module"
        else:
            plt.ylabel("Best modules error", fontsize=25)
            plt.title("Ensemble error vs Best module error", fontsize=25)
            context["result_name"] += "best_module"

        plt.xlim(xmin=0., xmax=context["graphics"]["scatter"]["xmax"])
        plt.ylim(ymin=0., ymax=context["graphics"]["scatter"]["ymax"])
        plt.xticks(np.linspace(0.0, context["graphics"]["scatter"]["xmax"], num=10), fontsize=8)
        plt.yticks(np.linspace(0.0, context["graphics"]["scatter"]["ymax"], num=10), fontsize=8)
        legend = plt.legend(loc="best")
        for t in legend.get_texts():
            t.set_fontsize('large')
        plt.savefig(
            context["classifiers"][context["classifier_list"][0]]["paths"]["results"] +
            context["pattern_kind"] + "_scatter_" + context["result_name"] + "." + context["graphics"]["extension"])

    ####################################################

    def best_choice(self, context, stats):
        f = open(context["classifiers"][context["classifier_list"][0]]["paths"]["results"]
                 + "configuration_evaluation.txt", "w")
        for name in [x for x in stats.measures.keys() if "evaluation" in stats.measures[x].keys()]:
            f.write(str(name) + "\n")
            for error, rms in zip(stats.measures[name]["selection"]["e"], stats.measures[name]["selection"]['rms']):
                f.write("Error=%d y rms=%f \n" % (error, rms))
            for neuron in stats.measures[name]["selection"]["neurons"]["hidden"]:
                f.write("With %s Neurons, amount %s" % (str(neuron),
                                                        str(stats.measures[name]["selection"]["neurons"][neuron][
                                                            "amount"])) + "\n")
            f.write("With names:\n")
            for name in stats.measures[name]["selection"]["names"]:
                f.write("%s," % name)
            f.write("\n")
            f.write("*********************\n\n")
        f.write("All selections:\n")
        for name in [x for x in stats.measures.keys() if "evaluation" in stats.measures[x].keys()]:
            for name in stats.measures[name]["selection"]["names"]:
                f.write("%s," % name)
        f.close()

    ####################################################
    ################################################################
    def barras(self, context, stats):
        import matplotlib.pyplot as plt
        import numpy as np

        efn = [.0, 0.066, 0.168, 0.029, 0.104, 0.210, 0.03, 0.202, 0.283, 0.037, 0.099, 0.033, 0.045, 0.028, 0.050,
               0.016]
        efp = [0.0, 0.0, 0.0, 0.0, 0.0, 0.042, 0.032, 0.058, 0.0, 0.0, 0.033, 0.0, 0.0, 0.025, 0.0, 0.0]
        efp_std = [0.0, 0.0, 0.0, 0.0, 0.0, 0.138, 0.122, 0.171, 0.0, 0.0, 0.125, 0.0, 0.0, 0.109, 0.0, 0.0]
        efn_std = [0.0, 0.179, 0.202, 0.099, 0.185, 0.201, 0.101, 0.206, 0.250, 0.104, 0.1065, 0.10, 0.118, 0.098,
                   0.126, 0.070]

        width = 0.7
        ind = np.arange(16)

        plt.rc('text', usetex=True)
        plt.rc('font', family='serif')

        p1 = plt.bar(ind, efn, width, color="k", yerr=efn_std, label=r"$E_{FN}$", ecolor="k", align='center')
        p2 = plt.bar(ind, efp, width, bottom=efn, color="r", yerr=efp_std, label=r"$E_{FP}$", ecolor="r",
                     align='center')

        plt.xlabel("Classifiers", fontsize=40)
        plt.ylabel("e", fontsize=40)
        legend = plt.legend(loc="best", prop={'size': 50})
        for t in legend.get_texts():
            t.set_fontsize('xx-large')
        plt.xticks(ind + width, (
            "", r"BPN$_1$", r"BPN$_2$", r"BPN$_3$", r"BPN$_4$", r"BPN$_5$", r"BPN$_6$", r"BPN$_7$", r"BPN$_8$",
            r"BPN$_{1-4}$", r"BPN$_{5-8}$", r"BPN$_{1-8}$", r"SMV$_{1-4}$", r"SMV$_{5-8}$", r"SMV$_{1-8}$",
            r"SMV$_{1+3}$"),
                   fontsize=32, rotation=30, ha='right')
        plt.yticks(np.arange(0.0, 0.501, 0.05))
        ymin, ymax = plt.ylim()
        plt.ylim(ymin=0.0, ymax=0.501)
        plt.tight_layout()
        plt.show()
